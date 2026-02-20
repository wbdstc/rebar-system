"""
Rebar System - Python Flask Backend
V12.0 企业级架构版

功能特性:
- SQLAlchemy 数据库支持 (MySQL)
- MinIO 对象存储
- 完整的 API 接口
- CORS 跨域支持

启动方式: python app.py
"""

import os
import io
import uuid
import tempfile
from datetime import datetime
from flask import Flask, request, jsonify, send_file, redirect
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
import requests as http_requests

# 本地工具模块
from utils.geometry import generate_hoop_path
from utils.ocr_helper import (
    process_ocr_image, 
    parse_pingfa, 
    get_design_total,
    check_compliance
)
from utils.vlm_service import parse_cad_image, verify_material

# Excel 处理
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# MinIO 客户端
try:
    from minio import Minio
    from minio.error import S3Error
    MINIO_AVAILABLE = True
except ImportError:
    MINIO_AVAILABLE = False
    print("[警告] minio 未安装，文件将存储在本地")

# ===========================================
# 应用配置
# ===========================================

app = Flask(__name__, static_folder='public', static_url_path='')
CORS(app, resources={r"/api/*": {"origins": "*"}})

# 数据库配置
DB_CONFIG = {
    'host': os.environ.get('MYSQL_HOST', 'localhost'),
    'port': int(os.environ.get('MYSQL_PORT', 3306)),
    'user': os.environ.get('MYSQL_USER', 'root'),
    'password': os.environ.get('MYSQL_PASSWORD', 'root'),
    'database': 'rebar_system'
}

app.config['SQLALCHEMY_DATABASE_URI'] = (
    f"mysql+pymysql://{DB_CONFIG['user']}:{DB_CONFIG['password']}"
    f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
    f"?charset=utf8mb4"
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_recycle': 3600,
    'pool_pre_ping': True
}

# 初始化数据库
db = SQLAlchemy(app)

# MinIO 配置
MINIO_CONFIG = {
    'endpoint': os.environ.get('MINIO_ENDPOINT', 'localhost:9000'),
    'access_key': os.environ.get('MINIO_ACCESS_KEY', 'minioadmin'),
    'secret_key': os.environ.get('MINIO_SECRET_KEY', 'minioadmin'),
    'secure': os.environ.get('MINIO_SECURE', 'false').lower() == 'true',
    'bucket': os.environ.get('MINIO_BUCKET', 'rebar-images')
}

# MinIO 客户端实例
minio_client = None
if MINIO_AVAILABLE:
    try:
        minio_client = Minio(
            MINIO_CONFIG['endpoint'],
            access_key=MINIO_CONFIG['access_key'],
            secret_key=MINIO_CONFIG['secret_key'],
            secure=MINIO_CONFIG['secure']
        )
        # 确保 bucket 存在
        if not minio_client.bucket_exists(MINIO_CONFIG['bucket']):
            minio_client.make_bucket(MINIO_CONFIG['bucket'])
            print(f"[MinIO] 创建 bucket: {MINIO_CONFIG['bucket']}")
        print(f"[MinIO] 连接成功: {MINIO_CONFIG['endpoint']}")
    except Exception as e:
        print(f"[警告] MinIO 连接失败: {e}")
        minio_client = None

# Roboflow API 配置
API_KEY = "TJD6y13Dru57zUSoA0D1"

MODELS = {
    'spacing': "https://detect.roboflow.com/rebar-4y6jc-vrqiw/3",
    'counting': "https://detect.roboflow.com/rebar-9zzhq-zm30m/1"
}

UPLOAD_FOLDER = tempfile.gettempdir()

# ===========================================
# 数据模型
# ===========================================

class InspectionRecord(db.Model):
    """检测记录表"""
    __tablename__ = 'inspection_records'
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    record_id = db.Column(db.String(50), unique=True, nullable=False)
    inspection_type = db.Column(db.String(20), nullable=False)
    project_name = db.Column(db.String(100))
    location = db.Column(db.String(200))
    column_id = db.Column(db.String(20))
    section_width = db.Column(db.Integer)
    section_height = db.Column(db.Integer)
    detected_count = db.Column(db.Integer)
    design_total = db.Column(db.Integer)
    compliance_status = db.Column(db.String(20))
    compliance_message = db.Column(db.Text)
    rebar_config = db.Column(db.JSON)
    predictions = db.Column(db.JSON)
    hoop_path = db.Column(db.JSON)
    image_url = db.Column(db.String(500))
    result_image_url = db.Column(db.String(500))
    inspector = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    def to_dict(self):
        return {
            'id': self.id,
            'record_id': self.record_id,
            'inspection_type': self.inspection_type,
            'project_name': self.project_name,
            'location': self.location,
            'column_id': self.column_id,
            'section_size': [self.section_width, self.section_height] if self.section_width else None,
            'detected_count': self.detected_count,
            'design_total': self.design_total,
            'compliance_status': self.compliance_status,
            'compliance_message': self.compliance_message,
            'rebar_config': self.rebar_config,
            'image_url': self.image_url,
            'inspector': self.inspector,
            'created_at': self.created_at.isoformat() if self.created_at else None,
        }

# ===========================================
# MinIO 文件操作
# ===========================================

def upload_to_minio(file_data: bytes, filename: str, content_type: str = 'image/jpeg') -> str:
    """
    上传文件到 MinIO
    
    Args:
        file_data: 文件二进制数据
        filename: 文件名
        content_type: MIME 类型
        
    Returns:
        文件的访问 URL
    """
    if not minio_client:
        # MinIO 不可用，保存到本地
        local_path = os.path.join(UPLOAD_FOLDER, filename)
        with open(local_path, 'wb') as f:
            f.write(file_data)
        return f"file://{local_path}"
    
    try:
        # 生成唯一文件名
        ext = os.path.splitext(filename)[1]
        unique_name = f"{datetime.now().strftime('%Y%m%d')}/{uuid.uuid4().hex}{ext}"
        
        # 上传到 MinIO
        minio_client.put_object(
            MINIO_CONFIG['bucket'],
            unique_name,
            io.BytesIO(file_data),
            length=len(file_data),
            content_type=content_type
        )
        
        # 返回访问 URL
        protocol = 'https' if MINIO_CONFIG['secure'] else 'http'
        url = f"{protocol}://{MINIO_CONFIG['endpoint']}/{MINIO_CONFIG['bucket']}/{unique_name}"
        
        print(f"[MinIO] 上传成功: {unique_name}")
        return url
        
    except S3Error as e:
        print(f"[MinIO] 上传失败: {e}")
        # 回退到本地存储
        local_path = os.path.join(UPLOAD_FOLDER, filename)
        with open(local_path, 'wb') as f:
            f.write(file_data)
        return f"file://{local_path}"


def get_from_minio(object_name: str) -> bytes:
    """从 MinIO 获取文件"""
    if not minio_client:
        return None
    
    try:
        response = minio_client.get_object(MINIO_CONFIG['bucket'], object_name)
        return response.read()
    except S3Error as e:
        print(f"[MinIO] 获取失败: {e}")
        return None

# ===========================================
# API 路由 - 静态页面
# ===========================================

@app.route('/')
def index():
    """默认首页重定向"""
    return redirect('/portal.html')

# ===========================================
# API 路由 - 检测接口
# ===========================================

@app.route('/analyze', methods=['POST'])
@app.route('/api/analyze', methods=['POST'])
def analyze():
    """统一分析接口"""
    if 'image' not in request.files:
        return jsonify({"error": "No image uploaded"}), 400
    
    image_file = request.files['image']
    if image_file.filename == '':
        return jsonify({"error": "No image selected"}), 400
    
    mode = request.args.get('mode', 'spacing')
    conf = request.args.get('conf', 40, type=int)
    overlap = request.args.get('overlap', 40, type=int)
    
    # 间距检测扩展参数
    component_type = request.args.get('component_type', '')  # slab_wall | beam_column
    pixel_per_mm = request.args.get('pixel_per_mm', 0, type=float)
    target_spacing = request.args.get('target_spacing', 150, type=float)
    target_spacing_dense = request.args.get('target_spacing_dense', 100, type=float)
    target_spacing_sparse = request.args.get('target_spacing_sparse', 200, type=float)
    tolerance = request.args.get('tolerance', 20, type=float)
    
    print(f"收到请求 | 模式: {mode} | 置信度: {conf} | 重叠: {overlap} | 构件: {component_type}")
    
    # 读取图片数据
    image_data = image_file.read()
    image_file.seek(0)
    
    # 上传到 MinIO
    image_url = upload_to_minio(
        image_data, 
        f"input_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
    )
    
    # 保存临时文件用于 AI 分析
    temp_path = os.path.join(UPLOAD_FOLDER, f"rebar_{uuid.uuid4().hex}.jpg")
    with open(temp_path, 'wb') as f:
        f.write(image_data)
    
    try:
        if mode == 'column':
            result = process_column_detection(temp_path, conf, overlap)
        else:
            result = call_roboflow_api(temp_path, mode, conf, overlap)
        
        # 间距合规性检查
        if mode == 'spacing' and component_type and pixel_per_mm > 0:
            spacings = process_spacing_check(
                predictions=result.get('predictions', []),
                component_type=component_type,
                pixel_per_mm=pixel_per_mm,
                target_spacing=target_spacing,
                target_spacing_dense=target_spacing_dense,
                target_spacing_sparse=target_spacing_sparse,
                tolerance=tolerance
            )
            result['spacings'] = spacings
        
        # 添加图片 URL
        result['image_url'] = image_url
        
        return jsonify(result)
    
    except Exception as e:
        print(f"处理失败: {str(e)}")
        return jsonify({"error": str(e)}), 500
    
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def call_roboflow_api(image_path: str, mode: str, conf: int, overlap: int) -> dict:
    """调用 Roboflow API"""
    target_url = MODELS.get(mode, MODELS['spacing'])
    
    with open(image_path, 'rb') as f:
        response = http_requests.post(
            target_url,
            params={
                "api_key": API_KEY,
                "confidence": conf,
                "overlap": overlap
            },
            files={"file": f}
        )
    
    if response.status_code == 200:
        data = response.json()
        print(f"识别成功，目标数: {len(data.get('predictions', []))}")
        return data
    else:
        raise Exception(f"Roboflow API 调用失败: {response.status_code}")


def process_spacing_check(predictions, component_type, pixel_per_mm,
                          target_spacing=150, target_spacing_dense=100,
                          target_spacing_sparse=200, tolerance=20):
    """
    间距合规性检查
    
    根据构件类型判定每段相邻钢筋间距是否合格：
    - 板/墙 (slab_wall): 单一设计间距判定
    - 梁/柱 (beam_column): 加密区 + 非加密区双间距判定
    
    Returns:
        list[dict]: 每段间距的检查结果，含坐标、距离、状态、颜色
    """
    if not predictions or len(predictions) < 2 or pixel_per_mm <= 0:
        return []
    
    # 提取中心坐标
    centers = [(p['x'], p['y']) for p in predictions]
    
    # 根据坐标分布判断钢筋排列主方向
    xs = [c[0] for c in centers]
    ys = [c[1] for c in centers]
    x_range = max(xs) - min(xs)
    y_range = max(ys) - min(ys)
    
    # 按主方向排序
    if x_range >= y_range:
        centers.sort(key=lambda c: c[0])
    else:
        centers.sort(key=lambda c: c[1])
    
    spacings = []
    for i in range(len(centers) - 1):
        sx, sy = centers[i]
        ex, ey = centers[i + 1]
        
        # 计算像素距离并转换为 mm
        px_dist = ((ex - sx) ** 2 + (ey - sy) ** 2) ** 0.5
        mm_dist = px_dist / pixel_per_mm
        
        spacing_info = {
            'index': i,
            'start': {'x': round(sx, 1), 'y': round(sy, 1)},
            'end':   {'x': round(ex, 1), 'y': round(ey, 1)},
            'px_distance': round(px_dist, 1),
            'mm_distance': round(mm_dist, 1),
        }
        
        if component_type == 'slab_wall':
            # 板/墙：单一设计间距
            if abs(mm_dist - target_spacing) <= tolerance:
                spacing_info['status'] = 'pass'
                spacing_info['color'] = '#00e676'   # 绿色
                spacing_info['label'] = '合格'
            else:
                spacing_info['status'] = 'fail'
                spacing_info['color'] = '#ff1744'   # 红色
                spacing_info['label'] = '不合格'
        
        elif component_type == 'beam_column':
            # 梁/柱：双间距判定
            diff_dense  = abs(mm_dist - target_spacing_dense)
            diff_sparse = abs(mm_dist - target_spacing_sparse)
            
            if diff_dense <= tolerance:
                spacing_info['status'] = 'pass_dense'
                spacing_info['color'] = '#00e5ff'   # 青色 Cyan - 加密区合格
                spacing_info['label'] = '加密区合格'
            elif diff_sparse <= tolerance:
                spacing_info['status'] = 'pass_sparse'
                spacing_info['color'] = '#00e676'   # 绿色 - 非加密区合格
                spacing_info['label'] = '非加密区合格'
            else:
                spacing_info['status'] = 'fail'
                spacing_info['color'] = '#ff1744'   # 红色 - 不合格
                spacing_info['label'] = '不合格'
        
        spacings.append(spacing_info)
    
    # 统计
    total = len(spacings)
    passed = sum(1 for s in spacings if s['status'] != 'fail')
    print(f"间距检查完成 | 总计: {total} | 合格: {passed} | 不合格: {total - passed}")
    
    return spacings


def process_column_detection(image_path: str, conf: int, overlap: int) -> dict:
    """V11.0 柱截面检测"""
    ai_result = call_roboflow_api(image_path, 'counting', conf, overlap)
    predictions = ai_result.get('predictions', [])
    detected_count = len(predictions)
    
    hoop_data = generate_hoop_path(predictions)
    
    return {
        "predictions": predictions,
        "detected_count": detected_count,
        "hoop_path": hoop_data["outer_hoop"],
        "inner_ties": hoop_data["inner_ties"],
        "image": ai_result.get("image", {}),
        "time": ai_result.get("time", 0)
    }

# ===========================================
# API 路由 - CAD 图纸智能解析
# ===========================================

@app.route('/api/parse_cad', methods=['POST'])
def parse_cad():
    """CAD 图纸智能解析接口 —— 根据构件类型调用大模型提取平法参数"""
    if 'image' not in request.files:
        return jsonify({"success": False, "error": "未上传 CAD 截图"}), 400

    image_file = request.files['image']
    if image_file.filename == '':
        return jsonify({"success": False, "error": "未选择文件"}), 400

    component_type = request.form.get('component_type', 'column')
    image_data = image_file.read()
    print(f"[CAD 解析] 构件={component_type}, 图片={image_file.filename}, 大小={len(image_data)} bytes")

    try:
        result = parse_cad_image(image_data, component_type)
        return jsonify(result)
    except Exception as e:
        print(f"[CAD 解析] 失败: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

# ===========================================
# API 路由 - 原材微观核验（轧印识别）
# ===========================================

@app.route('/api/verify_material', methods=['POST'])
def verify_material_api():
    """原材微观核验接口 —— 识别钢筋表面轧印"""
    if 'image' not in request.files:
        return jsonify({"success": False, "error": "未上传图片"}), 400

    image_file = request.files['image']
    if image_file.filename == '':
        return jsonify({"success": False, "error": "未选择文件"}), 400

    image_data = image_file.read()
    print(f"[原材核验] 图片={image_file.filename}, 大小={len(image_data)} bytes")

    try:
        result = verify_material(image_data)
        return jsonify(result)
    except Exception as e:
        print(f"[原材核验] 失败: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

# ===========================================
# API 路由 - OCR 接口
# ===========================================

@app.route('/ocr', methods=['POST'])
@app.route('/api/ocr', methods=['POST'])
def ocr_recognize():
    """OCR 识别接口"""
    if 'image' not in request.files:
        return jsonify({"error": "No image uploaded"}), 400
    
    image_file = request.files['image']
    image_data = image_file.read()
    
    try:
        result = process_ocr_image(image_data)
        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ===========================================
# API 路由 - 合规性检查
# ===========================================

@app.route('/check_compliance', methods=['POST'])
@app.route('/api/check_compliance', methods=['POST'])
def check_compliance_api():
    """合规性检查接口"""
    data = request.get_json()
    detected_count = data.get('detected_count', 0)
    design_total = data.get('design_total', 0)
    
    result = check_compliance(detected_count, design_total)
    return jsonify(result)

# ===========================================
# API 路由 - 检测记录 CRUD
# ===========================================

@app.route('/api/records', methods=['GET'])
def get_records():
    """获取检测记录列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    inspection_type = request.args.get('type')
    
    query = InspectionRecord.query.order_by(InspectionRecord.created_at.desc())
    
    if inspection_type:
        query = query.filter_by(inspection_type=inspection_type)
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'records': [r.to_dict() for r in pagination.items],
        'total': pagination.total,
        'page': page,
        'per_page': per_page,
        'pages': pagination.pages
    })


@app.route('/api/records', methods=['POST'])
def create_record():
    """创建检测记录"""
    data = request.get_json()
    
    record = InspectionRecord(
        record_id=f"IR{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4].upper()}",
        inspection_type=data.get('inspection_type', 'spacing'),
        project_name=data.get('project_name'),
        location=data.get('location'),
        column_id=data.get('column_id'),
        section_width=data.get('section_size', [None, None])[0] if data.get('section_size') else None,
        section_height=data.get('section_size', [None, None])[1] if data.get('section_size') else None,
        detected_count=data.get('detected_count'),
        design_total=data.get('design_total'),
        compliance_status=data.get('compliance', {}).get('status'),
        compliance_message=data.get('compliance', {}).get('message'),
        rebar_config=data.get('rebar_config'),
        predictions=data.get('predictions'),
        hoop_path=data.get('hoop_path'),
        image_url=data.get('image_url'),
        inspector=data.get('inspector')
    )
    
    db.session.add(record)
    db.session.commit()
    
    return jsonify({'success': True, 'record': record.to_dict()}), 201


@app.route('/api/records/<int:id>', methods=['GET'])
def get_record(id):
    """获取单条记录"""
    record = InspectionRecord.query.get_or_404(id)
    return jsonify(record.to_dict())


@app.route('/api/records/<int:id>', methods=['DELETE'])
def delete_record(id):
    """删除记录"""
    record = InspectionRecord.query.get_or_404(id)
    db.session.delete(record)
    db.session.commit()
    return jsonify({'success': True})

# ===========================================
# API 路由 - Excel 导出
# ===========================================

@app.route('/export_excel', methods=['POST'])
@app.route('/api/export_excel', methods=['POST'])
def export_excel():
    """导出 Excel 报表"""
    if not EXCEL_AVAILABLE:
        return jsonify({"error": "openpyxl 未安装"}), 500
    
    data = request.get_json()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "柱截面检测报告"
    
    # 样式定义
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 标题
    ws.merge_cells('A1:F1')
    ws['A1'] = "钢筋工程智能管控平台 - 柱截面检测报告"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    
    # 基本信息
    ws['A3'] = "柱号"
    ws['B3'] = data.get('column_id', 'N/A')
    ws['C3'] = "检测时间"
    ws['D3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    ws['A4'] = "截面尺寸"
    section = data.get('section_size')
    ws['B4'] = f"{section[0]}×{section[1]} mm" if section else "N/A"
    
    # 检测结果
    ws['A6'] = "检测结果汇总"
    ws['A6'].font = Font(bold=True, size=12)
    
    headers = ['A7', 'B7', 'C7']
    for cell, value in zip(headers, ['项目', '数值', '说明']):
        ws[cell] = value
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].alignment = center_align
        ws[cell].border = thin_border
    
    ws['A8'], ws['B8'], ws['C8'] = "AI 检测数量", data.get('detected_count', 0), "纵筋根数"
    ws['A9'], ws['B9'], ws['C9'] = "设计数量", data.get('design_total', 0), "图纸要求"
    
    compliance = data.get('compliance', {})
    ws['A10'] = "合规性判定"
    ws['B10'] = compliance.get('status', 'N/A')
    ws['C10'] = compliance.get('message', '')
    
    status = compliance.get('status', '')
    color_map = {'PASS': '008000', 'FAIL': 'FF0000', 'WARNING': 'FF8C00'}
    if status in color_map:
        ws['B10'].font = Font(color=color_map[status], bold=True)
    
    # 列宽
    for col, width in [('A', 15), ('B', 15), ('C', 25), ('D', 25)]:
        ws.column_dimensions[col].width = width
    
    # 保存
    filename = f"column_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    temp_path = os.path.join(UPLOAD_FOLDER, filename)
    wb.save(temp_path)
    
    return send_file(
        temp_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ===========================================
# 健康检查
# ===========================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查接口"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'database': 'connected' if db else 'disconnected',
        'minio': 'connected' if minio_client else 'disconnected',
        'excel': 'available' if EXCEL_AVAILABLE else 'unavailable'
    })

# ===========================================
# 启动服务器
# ===========================================

if __name__ == '__main__':
    PORT = int(os.environ.get('PORT', 5000))
    
    print("=" * 60)
    print("钢筋工程智能管控平台 - V12.0 企业级架构版")
    print("=" * 60)
    print(f"\n✅ Flask 后端启动: http://localhost:{PORT}")
    print(f"📁 静态文件目录: {os.path.abspath('public')}")
    print(f"🗄️  数据库: MySQL ({DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']})")
    print(f"📦 MinIO: {'已连接' if minio_client else '未连接'}")
    print(f"\n🔧 API 端点:")
    print(f"   - POST /api/analyze?mode=spacing|counting|column")
    print(f"   - POST /api/parse_cad        ← CAD 图纸智能解析")
    print(f"   - POST /api/verify_material  ← 原材微观核验")
    print(f"   - POST /api/ocr")
    print(f"   - GET/POST /api/records")
    print(f"   - POST /api/export_excel")
    print(f"   - GET /api/health")
    print("\n" + "=" * 60)
    
    app.run(host='0.0.0.0', port=PORT, debug=True)
